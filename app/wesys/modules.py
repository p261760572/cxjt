# coding=utf-8
import datetime
from decimal import Decimal
import json
import os
import re
import zipfile
import copy

import cx_Oracle
import xlrd
import xlutils.copy
from openpyxl import load_workbook,Workbook, worksheet

from app import utils, app
from app.gateway import hntl
from app.models import ElecInfo
from . import errors
from . import logger
from . import parser
from .exceptions import WesysException
from .exp import Expression
from .sql import Sql


class Module():
    def __init__(self):
        self._expr = Expression()
        self._sql = Sql()
        self._modules = {
            "select_one": self.select_one,
            "select_list": self.select_list,
            "select_page": self.select_page,
            "insert": self.insert,
            "update": self.update,
            "delete": self.delete,
            "copy": self.copy,
            "add": self.add,
            "callproc": self.callproc,
            "batch_select_one": self.batch_select_one,
            "batch_select_list": self.batch_select_list,
            "batch_insert": self.batch_insert,
            "batch_update": self.batch_update,
            "batch_delete": self.batch_delete,
            "batch_copy": self.batch_copy,
            "batch_callproc": self.batch_callproc,
            "export_txt": self.export_txt,
            "export_xls": self.export_xls,
            "import_txt": self.import_txt,
            "import_xls": self.import_xls,
            "zip": self.zip,
            "change_password": self.change_password,
            "hntl_gateway": self.hntl_gateway,
            "dynamic_sheet": self.dynamic_sheet
        }

    def reload(self):
        self._sql.reload()

    def get(self, name):
        return self._modules.get(name)

    def _select_list(self, ctx, sql_id, limit):
        sql, bind = self._sql.generate_sql(ctx, sql_id)
        rowcount, rows = ctx.conn.execute(sql, bind, limit)

        return rowcount, rows

    def _select_one(self, ctx, sql_id):
        sql, bind = self._sql.generate_sql(ctx, sql_id)
        rowcount, rows = ctx.conn.execute(sql, bind, limit=1)

        return rowcount, rows[0] if len(rows) > 0 else None

    def _update(self, ctx, sql_id):
        sql, bind = self._sql.generate_sql(ctx, sql_id)
        rowcount, _ = ctx.conn.execute(sql, bind)

        return rowcount

    def _insert(self, ctx, sql_id):
        return self._update(ctx, sql_id)

    def _delete(self, ctx, sql_id):
        return self._update(ctx, sql_id)

    def _callproc(self, ctx, procname, params):
        cur = ctx.conn.cursor()

        try:
            parameters = []
            for key in params:
                if key.startswith("#"):
                    parameters.append(cur.var(cx_Oracle.STRING))  # 全部做cx_Oracle.STRING处理
                elif key.startswith(":"):
                    parameters.append(ctx.get(key[1:]))
                else:
                    parameters.append(key)

            logger.error(procname)
            logger.error(parameters)
            copy_parameters = cur.callproc(procname, parameters)

        finally:
            cur.close()

        return copy_parameters

    def test_value(self, ctx, test):
        result = True
        if test:
            result = self._expr.calc(ctx, test)
            logger.info(test)
            logger.info(result)

        return result

    def _check_data_access(self, ctx, sql_id, request_data):
        sql, bind = self._sql.generate_sql(ctx, sql_id)

        # 去输出列
        m = re.search(r"[\s*']from[\s(]", sql)
        if m is None:
            raise WesysException(errors.ERROR_SQL_CONFIG)
        start = m.start() + 1

        total_sql = "select count(1) total " + sql[start:]

        _, rows = ctx.conn.execute(total_sql, bind)
        total = rows[0]["total"]

        return total

    def select_list(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        sql_id = params[0]
        key = "rows"
        rowcount = 1000
        if params_len > 1:
            key = params[1]

        if params_len > 2:
            rowcount = int(params[2])

        rowcount, rows = self._select_list(ctx, sql_id, rowcount)

        if key == "request":
            request_data["rows"] = rows
        if key.startswith("request."):
            request_data[key[8:]] = rows
        else:
            response_data[key] = rows

        return 0

    def select_page(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        sql_id = params[0]
        total_type = "fast"
        if params_len > 1:
            total_type = params[1]

        page = request_data.get("page")
        page_size = request_data.get("rows")

        if not str(page).isdigit():
            page = 1

        if not str(page_size).isdigit():
            page_size = 10

        sql, bind = self._sql.generate_sql(ctx, sql_id)

        if total_type == "fast":
            # 去输出列
            m = re.search(r"[\s*']from[\s(]", sql)
            if m is None:
                raise WesysException(errors.ERROR_SQL_CONFIG)
            start = m.start() + 1

            # 去order by
            m = re.search(r"[\s)]order\s+by\s", sql)
            end = len(sql)
            if m is not None:
                end = m.start() + 1

            total_sql = "select count(1) total " + sql[start:end]
        else:
            total_sql = "select count(1) total from (" + sql + ")"

        _, rows = ctx.conn.execute(total_sql, bind)
        total = rows[0]["total"]

        l = len(bind)
        rows_sql = "select * from (select rownum rn, a.* from (" + sql + ") a) where rn <= :b%d*:b%d and rn > (:b%d-1)*:b%d" % (
            l, l + 1, l + 2, l + 3)
        bind.append(page)
        bind.append(page_size)
        bind.append(page)
        bind.append(page_size)

        _, rows = ctx.conn.execute(rows_sql, bind)

        response_data["rows"] = rows
        response_data["total"] = total

        return 0

    def select_one(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        sql_id = params[0]
        key = "data"

        if params_len > 1:
            key = params[1]

        _, row = self._select_one(ctx, sql_id)

        if key == "request":
            data = request_data
        else:
            data = response_data.get(key)
            if data is None:
                data = {}
                response_data[key] = data

        if row is not None:
            data.update(row)

        return 0

    def insert(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        sql_id = params[0]
        data_sql_id = None
        key_sql_id = None

        if params_len > 1:
            data_sql_id = params[1]

        if params_len > 2:
            key_sql_id = params[2]

        if key_sql_id:
            _, row = self._select_one(ctx, key_sql_id)
            if row:
                request_data.update(row)

        rowcount = self._insert(ctx, sql_id)

        if data_sql_id:
            if self._check_data_access(ctx, data_sql_id, request_data) != rowcount:
                raise WesysException(errors.ERROR_DATA_ACCESS)

        return 0

    def update(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        sql_id = params[0]
        data_sql_id = None
        leastcount = 1

        if params_len > 1:
            data_sql_id = params[1]

        if params_len > 2:
            leastcount = int(params[2])

        rowcount = self._update(ctx, sql_id)

        if leastcount > rowcount:
            raise WesysException(errors.ERROR_NO_DATA_FOUND)

        if data_sql_id:
            if self._check_data_access(ctx, data_sql_id, request_data) != rowcount:
                raise WesysException(errors.ERROR_DATA_ACCESS)

        return 0

    def delete(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        sql_id = params[0]

        rowcount = self._delete(ctx, sql_id)

        if params_len > 1 and int(params[1]) > rowcount:
            raise WesysException(errors.ERROR_NO_DATA_FOUND)

        return 0

    def callproc(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        procname = params[0]
        argv = params[1:]

        logger.info(procname)
        logger.info(argv)

        copy_parameters = self._callproc(ctx, procname, argv)
        # output
        for i in range(0, len(argv)):
            if argv[i].startswith("#"):
                request_data[argv[i][1:]] = copy_parameters[i].strip()
        logger.info(argv)

        return 0

    def _copy_get(self, dest, request, response):
        dest_obj = None
        params = dest.split(".")
        if params[0] == "request":
            dest_obj = request
        elif params[0] == "response":
            dest_obj = response

        if len(params) > 1:
            dest_obj = dest_obj.get(params[1])

        return dest_obj

    def _copy_add(self, dest, request, response, default):
        dest_obj = None
        params = dest.split(".")
        if params[0] == "request":
            dest_obj = request
        elif params[0] == "response":
            dest_obj = response

        if len(params) > 1:
            temp = dest_obj.get(params[1])
            if temp is None:
                temp = default
                dest_obj[params[1]] = temp
            dest_obj = temp

        return dest_obj

    def copy(self, ctx, step, request_data, response_data):

        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 3:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        src = params[0]
        dest = params[1]
        key = params[2]
        new_key = key

        if params_len > 3:
            new_key = params[3]

        src_obj = self._copy_get(src, request_data, response_data)
        dest_obj = self._copy_get(dest, request_data, response_data)
        if dest_obj is None:
            dest_obj = self._copy_add(dest, request_data, response_data, {})

        if src_obj is None or dest_obj is None:
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        dest_obj[new_key] = src_obj.get(key)

        return 0

    def add(self, ctx, step, request_data, response_data):

        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len % 2 != 0:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        for i in range(0, params_len, 2):
            request_data[params[i]] = params[i + 1]

        return 0

    def batch_select_one(self, ctx, step, request_data, response_data):

        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in request_data:
            ctx.set(row)
            self.select_one(ctx, step, row, response_data)

        return 0

    def batch_select_list(self, ctx, step, request_data, response_data):

        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in request_data:
            ctx.set(row)
            self.select_list(ctx, step, row, response_data)

        return 0

    def batch_insert(self, ctx, step, request_data, response_data):

        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in request_data:
            ctx.set(row)
            self.insert(ctx, step, row, response_data)

        return 0

    def batch_update(self, ctx, step, request_data, response_data):

        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in request_data:
            ctx.set(row)
            self.update(ctx, step, row, response_data)

        return 0

    def batch_delete(self, ctx, step, request_data, response_data):

        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in request_data:
            ctx.set(row)
            self.delete(ctx, step, row, response_data)

        return 0

    def batch_callproc(self, ctx, step, request_data, response_data):

        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in request_data:
            ctx.set(row)
            self.callproc(ctx, step, row, response_data)

        return 0

    def batch_copy(self, ctx, step, request_data, response_data):
        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 3:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        src = params[0]
        dest = params[1]
        key = params[2]
        new_key = key

        if params_len > 3:
            new_key = params[3]

        src_obj = self._copy_get(src, request_data, response_data)
        dest_obj = self._copy_get(dest, request_data, response_data)

        if not isinstance(src_obj, dict):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        if not isinstance(dest_obj, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        for row in dest_obj:
            row[new_key] = src_obj.get(key)

        return 0

    def _generate_file(self, prefix, suffix):
        now = datetime.datetime.now()
        year = now.strftime("%Y")
        mon = now.strftime("%m")
        day = now.strftime("%d")

        serverid = utils.genereate_uuid().replace("+", "-").replace("/", "_")
        filename = prefix + serverid + suffix
        file_path = os.path.join(year, mon, day)
        pathname = os.path.join(app.config["EXPORT_FOLDER"], file_path)
        if not os.path.exists(pathname):
            os.makedirs(pathname)

        pathname = os.path.join(pathname, filename)
        os.mknod(pathname, mode=0o644)

        return pathname, os.path.join("/action/exports/", file_path, filename)

    def _to_string(self, obj):
        if obj is None:
            return ''
        return str(obj)

    def export_txt(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 3:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        sql_id = params[0]
        delim = params[1]
        prefix = params[2]
        if len(delim) == 0:
            delim = ","

        filename, url = self._generate_file(prefix, ".txt")
        f = open(filename, "w", encoding="GBK")

        sql, bind = self._sql.generate_sql(ctx, sql_id)
        cur = ctx.conn.cursor()
        try:
            logger.info(parser.generate_print_sql(sql, bind))
            cur.execute(sql, bind)
            while True:
                rows = cur.fetchmany(100)
                if len(rows) == 0:
                    break

                for row in rows:
                    f.write(delim.join(self._to_string(c) for c in row))
                    f.write("\r\n")
        finally:
            cur.close()

        f.close()

        temp = response_data.get("url")
        if temp is None:
            response_data["url"] = url
        elif isinstance(temp, list):
            temp.append(url)
        else:
            response_data["url"] = [temp, url]

        return 0

    def export_xls(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 6 or (params_len - 2) % 4 != 0:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        template = params[0]
        prefix = params[1]

        template_filename = os.path.join(app.config["TEMPLATE_FOLDER"], template)
        is_xls = template_filename.endswith(".xls")

        def _load_workbook(template_filename):
            if is_xls:
                rb = xlrd.open_workbook(template_filename, formatting_info=True, on_demand=True)
                wb = xlutils.copy.copy(rb)
            else:
                wb = load_workbook(template_filename)

            return wb

        def _get_sheet(wb, index):
            if is_xls:
                ws = wb.get_sheet(index)
            else:
                sheets = wb.get_sheet_names()
                ws = wb.get_sheet_by_name(sheets[index])

            return ws

        def _write_row(ws, r, c, row):
            if is_xls:
                i = c
                for col in row:
                    ws.write(r, i, col)
                    i += 1
            else:
                r += 1
                i = c + 1
                for col in row:
                    ws.cell(row=r, column=i, value=col)
                    i += 1

        if is_xls:
            filename, url = self._generate_file(prefix, ".xls")
        else:
            filename, url = self._generate_file(prefix, ".xlsx")

        wb = _load_workbook(template_filename)

        for i in range(2, params_len, 4):
            ws = _get_sheet(wb, int(params[i + 1]))
            r = int(params[i + 2])
            c = int(params[i + 3])

            sql, bind = self._sql.generate_sql(ctx, params[i])
            cur = ctx.conn.cursor()
            try:
                logger.info(parser.generate_print_sql(sql, bind))
                cur.execute(sql, bind)
                while True:
                    rows = cur.fetchmany(100)
                    if len(rows) == 0:
                        break

                    for row in rows:
                        _write_row(ws, r, c, row)
                        r += 1
            finally:
                cur.close()

        wb.save(filename)

        temp = response_data.get("url")
        if temp is None:
            response_data["url"] = url
        elif isinstance(temp, list):
            temp.append(url)
        else:
            response_data["url"] = [temp, url]

        return 0

    # def import_txt(self, ctx, step, request_data, response_data):
    #     if not self.test_value(ctx, step["test"]):
    #         return 0
    #
    #     params = step["param_list"].split(",")
    #     params_len = len(params)
    #
    #     if params_len < 4:
    #         raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)
    #
    #     batch_no = request_data.get(params[0])
    #     serverid = request_data.get(params[1])
    #     sql_id = params[2]
    #     delim = params[3]
    #
    #     if len(delim) == 0:
    #         delim = ","
    #
    #     elec_info = ElecInfo.query.filter_by(serverid=serverid).first()
    #     if elec_info is None:
    #         raise WesysException(errors.ERROR_NO_DATA_FOUND)
    #
    #     filename = os.path.join(app.config["UPLOAD_FOLDER"], elec_info.file_path)
    #     # insert sql
    #     sql, bind = self._sql.generate_sql(ctx, sql_id)
    #     bind_len = len(bind)
    #
    #     cur = ctx.conn.cursor()
    #     try:
    #         with open(filename, "r", encoding="GBK") as f:
    #             for line in f:
    #                 line = line.strip()
    #                 values = line.split(delim)
    #                 values.insert(0, batch_no)
    #                 values_len = len(values)
    #                 if values_len > bind_len:
    #                     values = values[0:bind_len]
    #                 else:
    #                     values = values + [None] * (bind_len - values_len)
    #                 logger.info(parser.generate_print_sql(sql, values))
    #                 cur.execute(sql, values)
    #     finally:
    #         cur.close()
    #
    #     f.close()
    #
    #     return 0

    # def import_xls(self, ctx, step, request_data, response_data):
    #     if not self.test_value(ctx, step["test"]):
    #         return 0
    #
    #     params = step["param_list"].split(",")
    #     params_len = len(params)
    #
    #     if params_len < 6 or (params_len - 2) % 4 != 0:
    #         raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)
    #
    #     batch_no = request_data.get(params[0])
    #     serverid = request_data.get(params[1])
    #
    #     elec_info = ElecInfo.query.filter_by(serverid=serverid).first()
    #     if elec_info is None:
    #         raise WesysException(errors.ERROR_NO_DATA_FOUND)
    #
    #     filename = os.path.join(app.config["UPLOAD_FOLDER"], elec_info.file_path)
    #     print(filename)
    #     rb = xlrd.open_workbook(filename)
    #
    #     for i in range(2, params_len, 4):
    #         rs = rb.sheet_by_index(int(params[i + 1]))
    #         r = int(params[i + 2])
    #         c = int(params[i + 3])
    #
    #         # insert sql
    #         sql, bind = self._sql.generate_sql(ctx, params[i])
    #         bind_len = len(bind)
    #
    #         cur = ctx.conn.cursor()
    #         try:
    #             for i in range(r, rs.nrows):
    #                 values = [batch_no]
    #                 for j in range(c, rs.ncols):
    #                     values.append(rs.cell(i, j).value)
    #                 values_len = len(values)
    #                 if values_len > bind_len:
    #                     values = values[0:bind_len]
    #                 else:
    #                     values = values + [None] * (bind_len - values_len)
    #
    #                 logger.info(parser.generate_print_sql(sql, values))
    #                 cur.execute(sql, values)
    #         finally:
    #             cur.close()
    #
    #     return 0

    def import_txt(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 9:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        serverid = params[0]
        delim = params[1]
        s_row = int(params[2])
        s_col = int(params[3])
        e_row = int(params[4])
        e_col = int(params[5])

        if serverid[0] == ':':
            serverid = request_data.get(params[0][1:])

        if len(delim) == 0:
            delim = ","

        elec_info = ElecInfo.query.filter_by(serverid=serverid).first()
        if elec_info is None:
            raise WesysException(errors.ERROR_NO_DATA_FOUND)

        filename = os.path.join(app.config["UPLOAD_FOLDER"], elec_info.file_path)

        sub_proc = copy.deepcopy(step)
        sub_proc['param_list'] = ','.join(params[6:])
        sub_proc["test"] = ''
        with open(filename, "r", encoding="GBK") as f:
            lines = f.readlines()
            for i, line in enumerate(lines[s_row:(len(lines)-e_row)]):
                line = line.strip()
                values = line.split(delim)
                line = delim.join(values[s_col:(len(values)-e_col)])
                request_data[params[7][1:]] = line
                request_data[params[8][1:]] = len(lines)
                request_data[params[9][1:]] = i + 1
                request_data[params[10][1:]] = delim
                self.callproc(ctx, sub_proc, request_data, response_data)

        f.close()

        return 0

    def import_xls(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 16 or (params_len-1) %15 !=0:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        serverid = params[0]
        sub_proc = copy.deepcopy(step)

        if serverid[0] == ':':
            serverid = request_data.get(params[0][1:])

        elec_info = ElecInfo.query.filter_by(serverid=serverid).first()
        if elec_info is None:
            raise WesysException(errors.ERROR_NO_DATA_FOUND)

        filename = os.path.join(app.config["UPLOAD_FOLDER"], elec_info.file_path)
        rb = xlrd.open_workbook(filename)

        for x in range(1, params_len, 15):
            rs = rb.sheet_by_index(int(params[x]))
            s_row = int(params[x+1])
            s_col = int(params[x+2])
            e_row = int(params[x+3])
            e_col = int(params[x+4])
            delim = params[x+5]

            if len(delim) == 0:
                delim = ','

            sub_proc['param_list'] = ','.join(params[x+6:x+6+9])
            sub_proc["test"] = ''

            for i in range(s_row, rs.nrows-e_row):
                values = []
                for j in range(s_col, rs.ncols-e_col):
                    values.append(rs.cell(i, j).value)
                    
                line = delim.join(map(str,values))
                request_data[params[x+7][1:]] = line
                request_data[params[x+8][1:]] = rs.nrows
                request_data[params[x+9][1:]] = i+1
                request_data[params[x+10][1:]] = delim
                self.callproc(ctx, sub_proc, request_data, response_data)

        return 0

    def zip(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        src = params[0]
        prefix = ""
        if params_len > 1:
            prefix = params[1]

        action_prefix = "/action/exports/"

        files = self._copy_get(src, request_data, response_data)
        zip_name, url = self._generate_file(prefix, ".zip")
        zf = zipfile.ZipFile(zip_name, "w", zipfile.zlib.DEFLATED)
        for f in files:
            if isinstance(f, dict):
                filename = os.path.join(app.config["UPLOAD_FOLDER"], f['file_path'])
                zf.write(filename, f['filename'])
            elif f.startswith(action_prefix):
                filename = os.path.join(app.config["EXPORT_FOLDER"], f[len(action_prefix):])
                zf.write(filename, os.path.basename(filename))
        zf.close()

        response_data["url"] = url

        return 0

    def change_password(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)

        userid = request_data.get("userid")
        login_pwd = request_data.get("login_pwd")

        if login_pwd:
            salt = utils.genereate_random_string(8, 32)
            ciphertext = utils.encrypt_password(login_pwd, salt)
            ctx.conn.execute("update userinfo set login_pwd=:login_pwd, salt=:salt where userid=:userid",
                             [ciphertext, salt, userid])

        return 0

    def hntl_gateway(self, ctx, step, request_data, response_data):
        if not self.test_value(ctx, step["test"]):
            return 0

        params = step["param_list"].split(",")
        params_len = len(params)
        action = params[0]

        if action == 'tms_quick_debug':
            r = hntl.tms_quick_debug(request_data)
            logger.debug(json.dumps(r, ensure_ascii=False))
            if r.get('errcode') == 0:
                request_data['gateway'] = r.get('data')
            else:
                raise WesysException(errors.ERROR_GATEWAY)
        else:
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        logger.debug(json.dumps(request_data, ensure_ascii=False))

        return 0


    def dynamic_sheet(self, ctx, step, request_data, response_data):
        if not isinstance(request_data, list):
            raise WesysException(errors.ERROR_MODULE_PARAM_ERROR)

        if not self.test_value(ctx, step["test"]):
            return 0

        if len(request_data) == 0:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        logger.debug(json.dumps(request_data, ensure_ascii=False))
        params = step["param_list"].split(",")
        params_len = len(params)

        if params_len < 1:
            raise WesysException(errors.ERROR_MODULE_CONFIG_ERROR)

        prefix = params[0]

        def _create_workbook():
            if is_xls:
                wb = xlwt.Workbook()
            else:
                wb = Workbook()
            return wb

        def _create_sheet(wb, sheet_name, index=None):
            if is_xls:
                ws = wb.add_sheet(sheet_name,cell_overwrite_ok=True)
            else:
                if index is not None:
                    ws = wb.create_sheet(sheet_name, 0)
                else:
                    ws = wb.create_sheet(sheet_name)
            return ws

        def _load_workbook(template_filename):
            if is_xls:
                rb = xlrd.open_workbook(template_filename, formatting_info=True, on_demand=True)
                wb = xlutils.copy.copy(rb)
            else:
                wb = load_workbook(template_filename)

            return wb

        def _get_sheet(wb, index):
            if is_xls:
                ws = wb.get_sheet(index)
            else:
                sheets = wb.get_sheet_names()
                ws = wb.get_sheet_by_name(sheets[index])
            return ws

        def _get_sheet_num(wb):
            if is_xls:
                s_n = len(wb.sheets())
            else:
                s_n = len(wb.get_sheet_names())

            return s_n

        def _remove_sheet(wb, index):
            ws = _get_sheet(wb, index)
            wb.remove_sheet(ws)

        def _set_sheet_name(ws, name):
            if is_xls:
                ws.name = name
            else:
                ws.title = name

        def _write_row(ws, r, c, row):
            if is_xls:
                i = c
                for col in row:
                    ws.write(r, i, col)
                    i += 1
            else:
                r += 1
                i = c + 1
                for col in row:
                    ws.cell(row=r, column=i, value=col)
                    i += 1

     ##   is_xls = template_filename.endswith(".xls")

        is_xls = False
        if is_xls:
            filename, url = self._generate_file(prefix, ".xls")
        else:
            filename, url = self._generate_file(prefix, ".xlsx")


   #     wb = _create_workbook()
        index = 0
        for loop in request_data:
            ctx.set(loop)
            sql, bind = self._sql.generate_sql(ctx, loop.get("loop_id"))
            cur = ctx.conn.cursor()

            template = loop.get("template")
            template_filename = os.path.join(app.config["TEMPLATE_FOLDER"], template)
            wb = _load_workbook(template_filename)

            try:
                logger.info(parser.generate_print_sql(sql, bind))
                cur.execute(sql, bind)
                while True:
                    sheet_loops = cur.fetchmany(100)
                    if len(sheet_loops) == 0:
                        break

                    for sheet_loop in sheet_loops:
                        r = loop.get("start_row")
                        c = loop.get("start_col")

                        cols = [d[0].lower() for d in cur.description]
                        values = []
                        for v in sheet_loop:
                            # 解决日期类型的序列化问题
                            if isinstance(v, datetime.datetime):
                                v = v.strftime("%Y-%m-%d %H:%M:%S")
                            elif isinstance(v, Decimal):
                                v = float(v)
                            values.append(v)
                        sheet_loop = dict(zip(cols, values))

                        ctx.update(sheet_loop)

                        sheet_name = sheet_loop.get("sheet_name")
                        #ws = _create_sheet(wb, sheet_name)
                        ws = _get_sheet(wb, index)
                        _set_sheet_name(ws, sheet_name)

                        sql2, bind2 = self._sql.generate_sql(ctx, loop.get("sql_id"))
                        cur2 = ctx.conn.cursor()
                        try:
                            logger.info(parser.generate_print_sql(sql2, bind2))
                            cur2.execute(sql2, bind2)
                            while True:
                                rows = cur2.fetchmany(100)
                                if len(rows) == 0:
                                    break

                                for row in rows:
                                    _write_row(ws, r, c, row)
                                    r += 1
                        finally:
                            cur2.close()
                            index += 1
            finally:
                cur.close()

        s_n = _get_sheet_num(wb)
        for i in range(index, s_n):
            _remove_sheet(wb, index)
        wb.save(filename)

        temp = response_data.get("url")
        if temp is None:
            response_data["url"] = url
        elif isinstance(temp, list):
            temp.append(url)
        else:
            response_data["url"] = [temp, url]

        return 0
"""
def write_oper_log(ctx, step, request_data, response_data):
    bind = []
    cur = ctx.sqlsession.conn.cursor()
    cur.execute("insert into oper_log(oper_time, client_ip, oper_id, action, error_msg) values (sysdate, :client_ip, :oper_id, :action, :error_msg)",  bind)
    cur.close()

"""
